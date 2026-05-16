"""One-time importer for the 2026 Expense Tracker workbook into
monthly_expenses. Walks each monthly sheet (Jan-Dec), picks up rows
from the FIXED / PAYROLL / OTHER sections, and inserts them with
inferred categories.

Run AFTER applying supabase/monthly_expenses.sql.

Usage:
  $env:SUPABASE_URL = "https://idddbbvotykfairirmwn.supabase.co"
  $env:SUPABASE_SERVICE_KEY = "<service role key>"
  python scripts/import_expense_tracker.py "C:/Users/bphet/Downloads/ROS_Expense_Tracker_2026.xlsx"

Each monthly sheet has this structure (header row varies slightly):
  R1:  "Jan 2026 — Expense Tracker"
  R2:  TOTAL EXPENSES formula | PAYROLL formula
  R4:  Date | Description | Amount | Category | Notes
  R5:  "FIXED BUSINESS EXPENSES" section header
  R6+: Fixed entries
  ~R75: PAYROLL section header
  ~R185: OTHER section header
  ~R267: TOTAL row

We detect section transitions by looking at column A text — if it
contains "PAYROLL" or "OTHER" we switch the inferred category. We
also respect the explicit Category column if it's set.
"""

import os
import sys
import json
from datetime import datetime
import openpyxl
import urllib.request

MONTHS = [
    'Jan 2026', 'Feb 2026', 'March 2026', 'April 2026',
    'May 2026', 'June 2026', 'July 2026', 'August 2026',
    'September 2026', 'October 2026', 'November 2026', 'December 2026',
]

IMPORT_TAG = 'expense_tracker_2026'

MONTH_NUM = {
    'Jan': '01', 'Feb': '02', 'March': '03', 'April': '04', 'May': '05', 'June': '06',
    'July': '07', 'August': '08', 'September': '09', 'October': '10', 'November': '11', 'December': '12',
}


def iso_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()
    if not s:
        return None
    if 'T' in s:
        s = s.split('T')[0]
    if ' ' in s:
        s = s.split(' ')[0]
    for fmt in ('%m/%d/%Y', '%Y-%m-%d', '%m/%d/%y'):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def num(v):
    if v is None or v == '':
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def is_section_header(text):
    if not text:
        return None
    upper = str(text).upper()
    if 'FIXED' in upper and 'EXPENSE' in upper:
        return 'Fixed'
    if 'PAYROLL' in upper:
        return 'Payroll'
    if upper.startswith('OTHER') or 'OTHER EXPENSE' in upper or 'OTHER BUSINESS' in upper:
        return 'Other'
    return None


def is_total_row(text):
    if not text:
        return False
    upper = str(text).upper().strip()
    return upper.startswith('TOTAL') or upper.startswith('GRAND TOTAL') or upper == 'TOTALS'


def parse_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    rows = []
    for sheet_name in MONTHS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        month_short = sheet_name.split(' ')[0]
        year = sheet_name.split(' ')[1]
        month_year = year + '-' + MONTH_NUM.get(month_short, '01')

        current_category = 'Fixed'  # Workbooks start with Fixed section first
        for r in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
            # Columns: A=Date, B=Description, C=Amount, D=Category, E=Notes
            col_a = clean(r[0])
            col_b = clean(r[1])
            col_c = num(r[2])
            col_d = clean(r[3])
            col_e = clean(r[4])

            # Section header detection (text in A spans the row)
            sec = is_section_header(col_a) or is_section_header(col_b)
            if sec:
                current_category = sec
                continue

            # Skip total/summary rows
            if is_total_row(col_a) or is_total_row(col_b):
                continue
            # Skip rows where everything is null
            if col_a is None and col_b is None and col_c is None and col_e is None:
                continue

            # Workbook is inconsistent: some sheets have Date in column A
            # and Description in B (Jan-April); others have Description
            # in A and Date in B (May+). Auto-detect by looking at which
            # column parses as a date.
            iso_a = iso_date(col_a)
            iso_b = iso_date(col_b)
            if iso_a and not iso_b:
                date = iso_a
                description = col_b
            elif iso_b and not iso_a:
                date = iso_b
                description = col_a
            elif iso_a and iso_b:
                # Both look like dates — use A as date, B as description fallback
                date = iso_a
                description = col_b
            else:
                # Neither is a date — default to 1st of month
                date = year + '-' + MONTH_NUM.get(month_short, '01') + '-01'
                description = col_b or col_a

            if not description:
                continue
            if col_c is None or col_c == 0:
                continue

            # Category: respect explicit column D, else use current section
            category = col_d if col_d in ('Fixed', 'Payroll', 'Other') else current_category

            rows.append({
                'date': date,
                'description': description,
                'amount': abs(col_c),
                'category': category,
                'notes': col_e,
                'vendor': None,
                'month_year': month_year,
                'imported_from': IMPORT_TAG,
            })
    return rows


def post_batch(url, key, batch):
    body = json.dumps(batch).encode('utf-8')
    req = urllib.request.Request(
        url + '/rest/v1/monthly_expenses',
        data=body,
        method='POST',
        headers={
            'apikey': key,
            'Authorization': 'Bearer ' + key,
            'Content-Type': 'application/json',
            'Prefer': 'return=minimal',
        },
    )
    try:
        with urllib.request.urlopen(req) as resp:
            if resp.status not in (200, 201, 204):
                raise RuntimeError(f'HTTP {resp.status}: {resp.read().decode()[:300]}')
    except urllib.error.HTTPError as e:
        body = e.read().decode()[:500]
        raise RuntimeError(f'HTTP {e.code}: {body}')


def fetch_existing_keys(url, key):
    """Returns set of (description, amount, month_year) tuples already in
    monthly_expenses, so re-running the import skips duplicates."""
    req = urllib.request.Request(
        url + '/rest/v1/monthly_expenses?select=description,amount,month_year',
        headers={'apikey': key, 'Authorization': 'Bearer ' + key},
    )
    with urllib.request.urlopen(req) as resp:
        rows = json.loads(resp.read())
        return {(r.get('description'), float(r.get('amount') or 0), r.get('month_year')) for r in rows}


def main():
    if len(sys.argv) < 2:
        print('Usage: python scripts/import_expense_tracker.py <path-to-xlsx>')
        sys.exit(1)
    path = sys.argv[1]
    supabase_url = os.environ.get('SUPABASE_URL')
    supabase_key = os.environ.get('SUPABASE_SERVICE_KEY')
    if not supabase_url or not supabase_key:
        print('SUPABASE_URL and SUPABASE_SERVICE_KEY env vars required')
        sys.exit(1)

    print(f'Parsing {path}...')
    rows = parse_workbook(path)
    print(f'Parsed {len(rows)} rows from workbook')

    existing = fetch_existing_keys(supabase_url, supabase_key)
    print(f'Existing entry count: {len(existing)}')

    # Deduplicate within batch AND against existing rows
    seen_in_batch = set()
    fresh = []
    for r in rows:
        key = (r['description'], r['amount'], r['month_year'])
        if key in existing or key in seen_in_batch:
            continue
        seen_in_batch.add(key)
        fresh.append(r)
    print(f'New rows to insert: {len(fresh)} (skipped {len(rows) - len(fresh)} duplicates)')

    if not fresh:
        print('Nothing to import.')
        return

    # Category counts
    by_cat = {}
    by_month = {}
    for r in fresh:
        by_cat[r['category']] = by_cat.get(r['category'], 0) + 1
        by_month[r['month_year']] = by_month.get(r['month_year'], 0) + 1
    print('By category:', by_cat)
    print('By month:', by_month)

    print('\nFirst 3 entries:')
    for r in fresh[:3]:
        print(' ', json.dumps(r, default=str))

    batch_size = 50
    total = 0
    for i in range(0, len(fresh), batch_size):
        batch = fresh[i:i + batch_size]
        post_batch(supabase_url, supabase_key, batch)
        total += len(batch)
        print(f'  posted {total}/{len(fresh)}')
    print(f'Done. {total} rows imported.')


if __name__ == '__main__':
    main()
