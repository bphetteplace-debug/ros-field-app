"""One-time importer for the Diamondback 2026 PM workbook into the
pm_schedule_entries Supabase table.

Run AFTER applying supabase/pm_schedule.sql.

Usage:
  SUPABASE_URL=https://idddbbvotykfairirmwn.supabase.co \\
  SUPABASE_SERVICE_KEY=<service-role-key> \\
  python scripts/import_diamondback_pm.py "C:/Users/bphet/Downloads/DiamondbackPM2026_FullYear_v4_MERGED (1).xlsx"

The script:
  - Parses each monthly sheet (Jan-Dec)
  - Skips section headers ("WEEK X · ..." rows) and header rows
  - Converts dates to ISO yyyy-mm-dd
  - Skips rows with no Location Name (template blanks)
  - Posts batches of 50 rows to /rest/v1/pm_schedule_entries
  - Tags each row imported_from='diamondback_2026_v4' so it's
    distinguishable from native entries later
"""

import os
import sys
import json
from datetime import datetime
import openpyxl
import urllib.request

MONTHS = [
    'January', 'February', 'March', 'April', 'May', 'June',
    'July', 'August', 'September', 'October', 'November', 'December'
]

CUSTOMER = 'Diamondback'
IMPORT_TAG = 'diamondback_2026_v4'

# Status normalization — collapse close variants to the canonical set
STATUS_MAP = {
    'Completed': 'Completed',
    'Scheduled PM': 'Scheduled PM',
    'Scheduled': 'Scheduled PM',
    'Needs Scheduling': 'Needs Scheduling',
    'Open': 'Open',
    'Delayed': 'Delayed',
    'Canceled': 'Canceled',
    'Cancelled': 'Canceled',
    'Flex Schedule': 'Flex Schedule',
}


def iso_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()
    if not s:
        return None
    # Accept 'YYYY-MM-DD HH:MM:SS' or 'YYYY-MM-DD'
    if 'T' in s:
        s = s.split('T')[0]
    if ' ' in s:
        s = s.split(' ')[0]
    # Quick validate
    try:
        datetime.strptime(s, '%Y-%m-%d')
        return s
    except ValueError:
        pass
    return None


def month_year_from_date(iso):
    if not iso:
        return None
    return iso[:7]


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    return s


def num(v):
    if v is None or v == '':
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def ticket_str(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    # Floats like '203712.0' → '203712'
    if s.endswith('.0'):
        s = s[:-2]
    return s


def parse_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    rows = []
    for month in MONTHS:
        if month not in wb.sheetnames:
            continue
        ws = wb[month]
        for r in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            # Columns: A=Location, B=Service Type, C=Area, D=Well Type,
            # E=Lat, F=Lng, G=Ticket#, H=Shut-In, I=Foreman, J=Status,
            # K=Notes, L=Assets, M=Date Completed
            loc = clean(r[0])
            if not loc:
                continue
            # Skip the duplicated header rows that appear inside some sheets
            if loc.lower() == 'location name':
                continue
            # Skip cells that are obviously section labels
            if loc.lower().startswith('week '):
                continue

            shut_in = iso_date(r[7])
            status_raw = clean(r[9]) or 'Needs Scheduling'
            status = STATUS_MAP.get(status_raw, status_raw)
            # If a date_completed is set but status is blank, infer Completed
            date_completed = iso_date(r[12])
            if date_completed and status == 'Needs Scheduling':
                status = 'Completed'

            rows.append({
                'customer': CUSTOMER,
                'location_name': loc,
                'service_type': clean(r[1]),
                'area': clean(r[2]),
                'well_type': clean(r[3]),
                'latitude': num(r[4]),
                'longitude': num(r[5]),
                'ticket_number': ticket_str(r[6]),
                'shut_in_date': shut_in,
                'foreman': clean(r[8]),
                'status': status,
                'notes': clean(r[10]),
                'assets': clean(r[11]),
                'date_completed': date_completed,
                'month_year': month_year_from_date(shut_in),
                'imported_from': IMPORT_TAG,
            })
    return rows


def post_batch(url, key, batch):
    body = json.dumps(batch).encode('utf-8')
    req = urllib.request.Request(
        url + '/rest/v1/pm_schedule_entries',
        data=body,
        method='POST',
        headers={
            'apikey': key,
            'Authorization': 'Bearer ' + key,
            'Content-Type': 'application/json',
            'Prefer': 'return=minimal',
        },
    )
    try:
        with urllib.request.urlopen(req) as resp:
            if resp.status not in (200, 201, 204):
                raise RuntimeError(f'HTTP {resp.status}: {resp.read().decode()[:300]}')
            return True
    except urllib.error.HTTPError as e:
        body = e.read().decode()[:500]
        raise RuntimeError(f'HTTP {e.code}: {body}')


def main():
    if len(sys.argv) < 2:
        print('Usage: python scripts/import_diamondback_pm.py <path-to-xlsx>')
        sys.exit(1)
    path = sys.argv[1]
    supabase_url = os.environ.get('SUPABASE_URL')
    supabase_key = os.environ.get('SUPABASE_SERVICE_KEY')
    if not supabase_url or not supabase_key:
        print('SUPABASE_URL and SUPABASE_SERVICE_KEY env vars required')
        sys.exit(1)

    print(f'Parsing {path}...')
    rows = parse_workbook(path)
    print(f'Parsed {len(rows)} entries')

    if not rows:
        print('No rows to import.')
        return

    # Sanity print first 3 rows
    print('First 3 entries:')
    for r in rows[:3]:
        print(' ', json.dumps(r, default=str))

    print(f'\nPosting {len(rows)} rows to {supabase_url} in batches of 50...')
    batch_size = 50
    total = 0
    for i in range(0, len(rows), batch_size):
        batch = rows[i:i + batch_size]
        post_batch(supabase_url, supabase_key, batch)
        total += len(batch)
        print(f'  posted {total}/{len(rows)}')
    print(f'Done. {total} rows imported.')


if __name__ == '__main__':
    main()
