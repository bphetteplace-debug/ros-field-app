"""One-time importer for the 2026 Workflow Tracker workbook into the
submissions table. Preserves original WKO numbers (Option B). Stuffs
payment-tracking fields onto submissions.data so the new 💵 Billing
admin tab picks them up automatically.

Usage:
  $env:SUPABASE_URL = "https://idddbbvotykfairirmwn.supabase.co"
  $env:SUPABASE_SERVICE_KEY = "<service role key>"
  python scripts/import_workflow_tracker.py "C:/Users/bphet/Downloads/2026_Work_Flow_Tracker.xlsx"

Each imported row:
  - template = 'service_call' or 'pm_flare_combustor' based on Work Type
  - work_order / pm_number = the original WKO# (preserves 1330, 1331, ...)
  - status = 'submitted'
  - date = ISO date from the workbook
  - customer_name, location_name, labor_hours, miles
  - total_revenue (top-level, for legacy fields)
  - data JSONB with:
      grandTotal, miles, labor_hours, techs (single-tech array)
      foreman (customer-side foreman from the workbook)
      paidDate, paidReference (parsed from the 'Paid' column)
      approvedDate (from 'Approved' column)
      paymentTerms (from DEFAULT_TERMS_BY_CUSTOMER)
      billable (from 'Billable' column — defaults true)
      dbWoNumber (the 'DB WO #' column)
      importedFrom = 'workflow_tracker_2026'
"""

import os
import re
import sys
import json
from datetime import datetime
import openpyxl
import urllib.request

MONTHS = [
    'January 2026', 'February 2026', 'March 2026', 'April 2026',
    'May 2026', 'June 2026', 'July 2026', 'August 2026',
    'September 2026', 'October 2026', 'November 2026', 'December 2026',
]

IMPORT_TAG = 'workflow_tracker_2026'

DEFAULT_TERMS = {
    'Diamondback': 'Net 60',
    'High Peak': 'Net 45',
    'ExTex': 'Net 30',
    'A8 Oilfield': 'Net 30',
    'KOS': 'Net 60',
    'Pristine Alliance': 'Net 30',
}

CUSTOMER_ALIASES = {
    'High Peak': 'High Peak',
    'Diamondback': 'Diamondback',
    'KOS': 'KOS',
    'ExTex': 'ExTex',
    'A8': 'A8 Oilfield',
    'A8 Oilfield': 'A8 Oilfield',
    'Pristine': 'Pristine Alliance',
    'Pristine Alliance': 'Pristine Alliance',
}


def iso_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()
    if not s:
        return None
    # Some cells have date + extra text (e.g. "3/27/2026 2000264804")
    parts = s.split()
    candidate = parts[0]
    # MM/DD/YYYY style
    for fmt in ('%m/%d/%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(candidate, fmt).date().isoformat()
        except ValueError:
            continue
    # T-separated ISO
    if 'T' in s:
        return s.split('T')[0]
    if ' ' in s:
        s = s.split(' ')[0]
        try:
            datetime.strptime(s, '%Y-%m-%d')
            return s
        except ValueError:
            return None
    return None


def parse_paid_cell(v):
    """The Paid column often has "3/27/2026 2000264804" — pull both."""
    if v is None:
        return None, None
    s = str(v).strip()
    if not s:
        return None, None
    parts = s.split(None, 1)
    paid_date = iso_date(parts[0])
    paid_ref = parts[1].strip() if len(parts) > 1 else None
    return paid_date, paid_ref


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def num(v):
    if v is None or v == '':
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def normalize_customer(v):
    s = clean(v)
    if not s:
        return None
    return CUSTOMER_ALIASES.get(s, s)


def classify_template(work_type):
    if not work_type:
        return 'service_call'
    s = str(work_type).strip().lower()
    if s in ('pm', 'preventive maintenance'):
        return 'pm_flare_combustor'
    return 'service_call'


def parse_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    rows = []
    for sheet_name in MONTHS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        # Header at row 4. Data starts at row 5.
        # Columns: A=Submitted, B=DB WO#, C=WKO#, D=Date, E=Billable,
        # F=Customer, G=Site, H=Work Type, I=Foreman, J=Tech, K=Notes,
        # L=Hours, M=Miles, N=Cost, O=Approved, P=Paid, Q=Needs Approval
        for r in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            wko_raw = r[2]
            customer = normalize_customer(r[5])
            site = clean(r[6])
            # Must have at least a customer + WKO# or some content
            if not wko_raw and not customer:
                continue
            try:
                wko = int(float(wko_raw)) if wko_raw not in (None, '') else None
            except (ValueError, TypeError):
                continue
            if wko is None:
                continue

            work_type = clean(r[7])
            template = classify_template(work_type)
            billable_cell = clean(r[4])
            is_billable = True
            if billable_cell and billable_cell.lower() in ('non', 'non-billable', 'no', 'false'):
                is_billable = False

            tech = clean(r[9])
            foreman = clean(r[8])
            notes = clean(r[10])
            hours = num(r[11]) or 0
            miles = num(r[12]) or 0
            cost = num(r[13]) or 0
            approved = iso_date(r[14])
            paid_date, paid_ref = parse_paid_cell(r[15])
            db_wo = clean(r[1])
            date = iso_date(r[3])

            data = {
                'jobType': 'PM' if template == 'pm_flare_combustor' else 'SC',
                'warrantyWork': False,
                'techs': [tech] if tech else [],
                'equipment': [],
                'parts': [],
                'miles': miles,
                'costPerMile': 1.50,
                'laborHours': hours,
                'hourlyRate': 115,
                'billableTechs': 1,
                'description': notes or '',
                'grandTotal': cost,
                'partsTotal': 0,
                'mileageTotal': 0,
                'laborTotal': cost,
                # Billing fields
                'dbWoNumber': db_wo,
                'foreman': foreman,
                'approvedDate': approved,
                'paidDate': paid_date,
                'paidReference': paid_ref,
                'paymentTerms': DEFAULT_TERMS.get(customer or '', 'Net 30'),
                'billable': is_billable,
                'workType': work_type,
                'importedFrom': IMPORT_TAG,
            }

            rows.append({
                'template': template,
                'customer_name': customer,
                'location_name': site,
                'date': date,
                'work_order': wko,
                'pm_number': wko,
                'labor_hours': hours,
                'miles': miles,
                'status': 'submitted',
                'summary': notes or '',
                'data': data,
            })
    return rows


def post_batch(url, key, batch):
    body = json.dumps(batch).encode('utf-8')
    req = urllib.request.Request(
        url + '/rest/v1/submissions',
        data=body,
        method='POST',
        headers={
            'apikey': key,
            'Authorization': 'Bearer ' + key,
            'Content-Type': 'application/json',
            'Prefer': 'return=minimal',
        },
    )
    try:
        with urllib.request.urlopen(req) as resp:
            if resp.status not in (200, 201, 204):
                raise RuntimeError(f'HTTP {resp.status}: {resp.read().decode()[:300]}')
            return True
    except urllib.error.HTTPError as e:
        body = e.read().decode()[:500]
        raise RuntimeError(f'HTTP {e.code}: {body}')


def fetch_existing_wkos(url, key):
    """Returns the set of numbers already used as work_order OR pm_number
    in submissions, so we don't trip the unique constraints on either."""
    req = urllib.request.Request(
        url + '/rest/v1/submissions?select=work_order,pm_number',
        headers={'apikey': key, 'Authorization': 'Bearer ' + key},
    )
    used = set()
    with urllib.request.urlopen(req) as resp:
        rows = json.loads(resp.read())
        for r in rows:
            if r.get('work_order') is not None:
                used.add(int(r['work_order']))
            if r.get('pm_number') is not None:
                used.add(int(r['pm_number']))
    return used


def main():
    if len(sys.argv) < 2:
        print('Usage: python scripts/import_workflow_tracker.py <path-to-xlsx>')
        sys.exit(1)
    path = sys.argv[1]
    supabase_url = os.environ.get('SUPABASE_URL')
    supabase_key = os.environ.get('SUPABASE_SERVICE_KEY')
    if not supabase_url or not supabase_key:
        print('SUPABASE_URL and SUPABASE_SERVICE_KEY env vars required')
        sys.exit(1)

    print(f'Parsing {path}...')
    rows = parse_workbook(path)
    print(f'Parsed {len(rows)} rows from workbook')

    existing = fetch_existing_wkos(supabase_url, supabase_key)
    print(f'Existing submissions WKO/pm count: {len(existing)}')
    # Skip any row whose WKO collides with existing AND dedupe within
    # the workbook itself (in case of typo'd duplicate rows).
    seen_in_batch = set()
    fresh = []
    for r in rows:
        n = r['work_order']
        if n in existing or n in seen_in_batch:
            continue
        seen_in_batch.add(n)
        fresh.append(r)
    print(f'New rows to insert: {len(fresh)} (skipped {len(rows) - len(fresh)} duplicates)')

    if not fresh:
        print('Nothing to import.')
        return

    print('First entry:')
    print(json.dumps(fresh[0], indent=2, default=str))

    batch_size = 50
    total = 0
    for i in range(0, len(fresh), batch_size):
        batch = fresh[i:i + batch_size]
        post_batch(supabase_url, supabase_key, batch)
        total += len(batch)
        print(f'  posted {total}/{len(fresh)}')
    print(f'Done. {total} rows imported.')


if __name__ == '__main__':
    main()
